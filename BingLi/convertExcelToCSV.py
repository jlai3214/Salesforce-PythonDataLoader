#!/usr/bin/python

__author__ = 'jeremylai'

from openpyxl import load_workbook
from sys import argv
import os
import subprocess

#### CONSTANTS ####
MAX_COL = 10


def writeNewRow(ws, targetFile, rowNum):
    for colNum in range(1, MAX_COL):
        if ws.cell(row=rowNum, column=colNum).value != None:
            targetFile.write(str(ws.cell(row=rowNum, column=colNum).value) + ",")
        else:
            targetFile.write(",")
    if ws.cell(row=rowNum, column=MAX_COL).value != None:
        targetFile.write(str(ws.cell(row=rowNum, column=MAX_COL).value))


def convertSheets(wb):
    # Do it for each tab/sheet in the Excel file
    for ws in wb:
        # New CSV file name
        fileName = ws.title + ".csv"
        targetFile = open(fileName, 'w')
        targetFile.truncate()

        # Write Header Info
        targetFile.write("Kenandy,Fall 2015,Fall 15,,,,,,,,,,\n")
        targetFile.write("execution,notes,")

        # Append to second row/Header info
        writeNewRow(ws, targetFile, 1)
        targetFile.write("\n")

        # Shift the other rows by two comma delimiters
        for rowNum in range(2, ws.max_row):
            targetFile.write(",,")
            writeNewRow(ws, targetFile, rowNum)
            targetFile.write("\n")

        # Will run Bing's csvConverter script, but requires a "CTRL + C" after it finishes
        # or it will never timeout
        p = subprocess.Popen("python csvConverter.py " + fileName, shell=True)


if __name__ == '__main__':
    if len(argv) < 2:
        print("ERROR: Missing arguement")
    elif len(argv) > 2:
        print("ERROR: Too many arguments")
    elif not os.path.exists(argv[1]):
        print("Cannot find file " + argv[1])
    else:
        wb = load_workbook(argv[1])
        convertSheets(wb)
